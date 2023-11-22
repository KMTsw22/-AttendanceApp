from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
import json
app = FastAPI()

# CORS 설정
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # 모든 origin에서의 요청을 허용하려면 "*"를 사용합니다.
    allow_credentials=True,
    allow_methods=["*"],  # 모든 HTTP 메서드를 허용합니다.
    allow_headers=["*"],  # 모든 헤더를 허용합니다.
)

@app.get('/')
def read_root():
    return {"message": "Hello, World"}

@app.post('/Student')
def get_infor():
    list_ = open_exel()
    return list_

@app.post('/item/{ID}')
def read_item(ID: str):
    print(ID)
    list_ = open_exel()
    list_second = []
    for i in list_:
        if i[2] == ID:
            list_second.append(i)

    return {"PhonePassList": list_second}

@app.post('/SendMessage/{Telephone}/{Name}')
def send_message(Telephone:str, Name: str):
    from sdk.api.message import Message
    from sdk.exceptions import CoolsmsException
    api_key = ''
    api_secret = ''
    params = dict()
    params['type'] = 'sms'  # Message type ( sms, lms, mms, ata )
    params['to'] = Telephone  # Recipients Number '01000000000,01000000001'
    params['from'] = ''  # Sender number
    params['text'] = f'{Name} 학생 출결'

    print('heloo')
    cool = Message(api_key, api_secret)
    print(cool)
    try:
        response = cool.send(params)
        print('메시지 보내기 성공')
        if "error_list" in response:
            print("Error List : %s" % response['error_list'])
            pass

        return {"Success Code": "Success"}
    except CoolsmsException as e:
        print("Error Code : %s" % e.code)
        print("Error Message : %s" % e.msg)
        return {"Success Code": "Fail"}
@app.post('/plus/{name}/{phone}/{id}/{pw}')
def plus_student(name:str, phone:str, id:str,pw:str):
    import openpyxl
    try:
        workbook = openpyxl.load_workbook('통합 문서1.xlsx')  # 파일명.xlsx에 실제 파일 이름을 입력하세요
        sheet = workbook.active  # 기본 시트를 선택하거나
        list_ = []
        max_row = sheet.max_row
        sheet.cell(max_row+1,1).value = name
        sheet.cell(max_row+1,2).value = phone
        sheet.cell(max_row+1,3).value = id
        sheet.cell(max_row+1,4).value = pw
        return {"Code": 'Success'}
    except:
        return {"Code": 'Fail'}

def get_infor():
    list_ = open_exel()
    return list_




def open_exel():
    import openpyxl
    workbook = openpyxl.load_workbook('통합 문서1.xlsx')  # 파일명.xlsx에 실제 파일 이름을 입력하세요
    sheet = workbook.active  # 기본 시트를 선택하거나
    list_ = []
    for i in range(1, sheet.max_row + 1):
        name = str(sheet.cell(i, 1).value).encode('utf-8')
        phone = str(sheet.cell(i, 2).value)
        phone_back = str(sheet.cell(i, 3).value)
        password = str(sheet.cell(i, 4).value)
        list_.append([name, phone, phone_back, password])
    return list_

